import os
import pandas as pd
import openpyxl
import fcsparser

# standard well dimensions
WELL_ROWS = 16
WELL_COLS = 24

# how far we search for the start of the well positions and sample lookup
ROWMAX = 100
COLMAX = 25

PLATE_PREFIXES = ['LCE', 'PRM']


def get_well_positions(meta):
    '''
    Extract well positions (e.g., A1, A2 etc.) from FCS file metadata,
    ordered by sorting locations.
    '''
    metalist = list(meta.keys())
    locnames = [item for item in metalist if item.startswith('INDEX SORTING LOCATIONS')]
    locnames = sorted(locnames, key=lambda s: int(s.split('_')[1]))

    wells = []
    for locname in locnames:
        loclist = meta[locname]
        locs = loclist.split(";")

        for loc in locs:
            if loc:
                x, y = loc.split(",")
                new_x = chr(int(x) + ord('A'))
                new_y = str(int(y) + 1)
                wells.append(new_x + new_y)

    return wells


def get_plate_and_sample_from_filepath(fcs_filepath):
    '''
    Extract plate and sample name from file names in format
    e.g., DDmonthYY_INX_samplename_platename.fcs where plate name
    starts with a prefix from PLATE_PREFIXES.
    '''
    filename = os.path.basename(fcs_filepath)
    filename = os.path.splitext(filename)[0]

    # Extract the plate name
    plate = None
    for prefix in PLATE_PREFIXES:
        plate_start_index = filename.find(prefix)
        if plate_start_index != -1:
            break

    if plate_start_index != -1:
        plate = filename[plate_start_index:]
    else:
        print('Could not find plate name in file name {}'.format(filename))

    # Extract the sample name
    sample_start_index = filename.find('INX_')
    if sample_start_index != -1:
        sample_name = filename[sample_start_index + len('INX_'):(plate_start_index - 1)]

    return plate, sample_name


def collate_fcs_files(fcs_files, upload_dir):
    '''
    Collate FCS files into a single dataframe, adding columns for plate and sample name.
    '''
    fcs_data = pd.DataFrame()

    for fcs_file in fcs_files:
        fcs_savepath = os.path.join(upload_dir, fcs_file.filename)
        meta, data = fcsparser.parse(fcs_savepath, meta_data_only=False, reformat_meta=True)
        data = data.sort_values('Time')
        data['well_position'] = get_well_positions(meta)

        plate, sample = get_plate_and_sample_from_filepath(fcs_savepath)
        data['plate'] = plate
        data['sample'] = sample

        fcs_data = pd.concat([fcs_data, data])

    return fcs_data


def get_sample_lookup(sheet, sample_start_cell):
    '''
    Extract sample lookup (colour index > sample name) from plate layout spreadsheet.
    '''
    sample_dict = {}

    current_cell = sample_start_cell
    tolerate_n_blank_rows = 2
    while(True):
        current_cell = sheet.cell(row=current_cell.row + 1, column=current_cell.column)
        colour = sheet.cell(row=current_cell.row, column=current_cell.column).fill.start_color.index

        if colour == '00000000':
            if tolerate_n_blank_rows == 0:
                break
            tolerate_n_blank_rows -= 1
            continue

        value = sheet.cell(current_cell.row, current_cell.column + 1).value
        sample_dict[colour] = value

    return sample_dict


def get_sample_list(sheet, sample_lookup, well_start_cell):
    '''
    iterate through the plate cells and extract
    sample name for each cell from the colour
    '''
    sample_list = []
    for row in range(0, WELL_ROWS):
        for col in range(1, WELL_COLS + 1):
            cell = sheet.cell(row=well_start_cell.row + row, column=well_start_cell.column + col)

            samplename = 'removed'
            if cell.fill.start_color.index != '00000000':
                cell_colour = cell.fill.start_color.index

                error_message = 'Cell colour {} in cell {} not found in sample lookup in sheet {}'.format(
                    cell_colour, cell.coordinate, sheet.title
                )
                assert cell_colour in sample_lookup, error_message
                samplename = sample_lookup[cell_colour]

            # get well ID
            well_row = sheet.cell(row=well_start_cell.row + row, column=well_start_cell.column).value
            well_col = sheet.cell(row=well_start_cell.row - 1, column=well_start_cell.column + col).value
            well_id = well_row + str(well_col)

            sample_list.append((well_id, samplename))

    return sample_list


def find_sample_start_cell(sheet):
    '''
    Find the start of the sample names in the spreadsheet.
    Return the cell.
    '''
    sample_start = None

    for col in range(1, COLMAX):
        for row in range(1, ROWMAX):
            cell = sheet.cell(row=row, column=col)
            if cell.value and str(cell.value).lower() in ['sort description', 'sort discription']:
                sample_start = cell
                break

    return sample_start


def find_well_start(sheet):
    '''
    Find the start of the well positions in the spreadsheet.
    This is the first cell containing just the value 'A' with
    a cell containing '1' above and to the right.
    '''
    well_start_cell = None
    for col in range(1, COLMAX):
        for row in range(1, ROWMAX):
            cell = sheet.cell(row=row, column=col)
            if cell.value == 'A':
                well_start_cell = cell
                # check that the cell above and to the right is 1
                if sheet.cell(row=row - 1, column=col + 1).value == 1:
                    return well_start_cell
    return None


def plate_to_samplesheet(xlsx_file):
    '''
    Convert plate layout spreadsheet to samplesheet.
    '''
    wb = openpyxl.load_workbook(xlsx_file)

    # iterate through all sheets
    full_samplesheet = pd.DataFrame()
    for sheet in wb:
        if not sheet.title.lower().startswith('lce'):
            print('Skipping sheet {}'.format(sheet.title))
            continue

        sample_start_cell = find_sample_start_cell(sheet)
        assert sample_start_cell, 'Could not find sample start cell in sheet {}'.format(sheet.title)

        well_start_cell = find_well_start(sheet)
        assert well_start_cell, 'Could not find well start cell in sheet {}'.format(sheet.title)

        sample_lookup = get_sample_lookup(sheet, sample_start_cell)
        sample_list = get_sample_list(sheet, sample_lookup, well_start_cell)
        samplesheet = pd.DataFrame(sample_list, columns=['well_position', 'sample'])
        samplesheet['plate'] = sheet.title

        full_samplesheet = pd.concat([full_samplesheet, samplesheet])

    # reorder columns
    full_samplesheet = full_samplesheet[['plate', 'well_position', 'sample']]
    return full_samplesheet


def load_excel_samplesheet(filepath):
    '''
    Load excel spreadsheet samplesheet and check it has a 'samples' sheet.
    Locate the first row and then load and return the data as a pandas dataframe.
    '''
    wb = openpyxl.load_workbook(filepath)
    assert sum([sheet.lower().startswith('sample') for sheet in wb.sheetnames]) == 1, 'Could not find samples sheet in template file'

    samples_sheet_name = [sheet for sheet in wb.sheetnames if sheet.lower().startswith('sample')][0]

    first_row, found_row = 1, False
    for row in wb[samples_sheet_name].iter_rows():
        if row[0].value and row[0].value.lower().startswith('plate'):
            found_row = True
            break
        first_row += 1
    assert found_row, 'Could not find header row in template file'

    samplesheet = pd.read_excel(filepath, sheet_name=samples_sheet_name, header=first_row - 1)

    assert 'Plate#' in samplesheet.columns, 'Could not find Plate# column in template file'
    assert 'Well position' in samplesheet.columns, 'Could not find Well position column in template file'

    return samplesheet


def merge_data_with_samplesheet(spreadsheet_filepath, fcs_file, template_sheet_filepath):
    '''
    Merges processed/uploaded data (may be sample sheet and/or fcs data).
    Will merge into a template if provided.
    '''
    fcs_data = pd.read_csv(fcs_file, sep='\t') if fcs_file else None
    is_xlsx = spreadsheet_filepath.endswith('.xlsx')
    merged_data = pd.DataFrame()

    if template_sheet_filepath and is_xlsx:
        raise Exception('Cannot merge template sheet with xlsx file. Please upload a tsv file with plate, sample and well positions.')
    elif template_sheet_filepath:
        template = load_excel_samplesheet(template_sheet_filepath)
        spreadsheet = pd.read_csv(spreadsheet_filepath, sep='\t')
        for plate in spreadsheet.plate.unique():
            plate_data = template.copy()
            plate_data['Plate#'] = plate
            plate_data = pd.merge(plate_data, spreadsheet,
                                  left_on=['Plate#', 'Well position'],
                                  right_on=['plate', 'well_position'], how='left')
            merged_data = pd.concat([merged_data, plate_data])
    elif not is_xlsx:
        merged_data = pd.read_csv(spreadsheet_filepath, sep='\t')

    if fcs_data is None:
        return merged_data

    if is_xlsx:
        spreadsheet = load_excel_samplesheet(spreadsheet_filepath)
        samples_colname = [col for col in spreadsheet.columns if col.lower() == 'sample' or col.lower() == 'sample name'][0]
        print(spreadsheet.head())
        print(fcs_data.head())
        return pd.merge(spreadsheet, fcs_data, 
                        left_on=['Plate#', 'Well position', samples_colname],
                        right_on=['plate', 'well_position', 'sample'], how='left')
    else:
        return pd.merge(merged_data, fcs_data, on=['plate', 'sample', 'well_position'], how='left')
